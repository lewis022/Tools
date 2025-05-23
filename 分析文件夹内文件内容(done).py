import os
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, numbers, PatternFill
import hashlib

FILE_TYPES = {
    "图片": {"png", "jpg", "jpeg", "gif", "bmp", "heif", "webp", "tiff", "heic", "arw"},
    "视频": {"mp4", "mov", "avi", "mkv", "flv", "m4v", "wmv", "mpeg"},
    "音频": {"mp3", "m4a", "wav", "flac", "aac", "ogg", "wma"},
    "文档": {"pdf", "doc", "docx", "ppt", "pptx", "txt", "md"},
    "表格": {"xlsx", "xls", "csv", "tsv"},
    "压缩包": {"zip", "rar", "7z", "tar", "gz"},
    "程序": {"exe", "dmg", "pkg", "app", "bat", "sh"},
    "代码": {"py", "java", "cpp", "c", "h", "html", "css", "js"}
}


def calculate_md5(file_path):
    """计算文件的MD5值"""
    hash_md5 = hashlib.md5()
    try:
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                if not chunk:
                    break
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except (IOError, PermissionError):
        return "计算失败"


def get_file_type(extension):
    ext = extension.lower().lstrip('.')
    for file_type, extensions in FILE_TYPES.items():
        if ext in extensions:
            return file_type
    return "其他" if ext else "未知"


def select_folder():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askdirectory()


def get_file_info(folder_path):
    file_list = []
    total_size = 0
    base_folder = os.path.normpath(folder_path)
    base_folder_name = os.path.basename(base_folder)
    max_depth = 0
    md5_dict = {}

    # 第一次遍历计算深度和大小
    for root, dirs, files in os.walk(folder_path):
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        files = [f for f in files if not f.startswith('.')]

        relative_path = os.path.relpath(root, base_folder)
        depth = len(relative_path.split(os.sep)) if relative_path != '.' else 0
        max_depth = max(max_depth, depth)

        for file in files:
            file_path = os.path.join(root, file)
            total_size += os.path.getsize(file_path)

    # 第二次遍历收集数据
    for root, dirs, files in os.walk(folder_path):
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        files = [f for f in files if not f.startswith('.')]

        for file in files:
            file_path = os.path.join(root, file)
            file_stat = os.stat(file_path)

            filename, ext = os.path.splitext(file)
            relative_path = os.path.relpath(root, base_folder)
            dir_components = relative_path.split(os.sep) if relative_path != '.' else []

            dir_levels = [base_folder_name] + dir_components
            file_size_mb = round(file_stat.st_size / (1024 * 1024), 2)
            size_percent = round(file_stat.st_size / total_size, 4) if total_size > 0 else 0
            md5_value = calculate_md5(file_path)

            # 记录MD5值出现的次数
            if md5_value in md5_dict:
                md5_dict[md5_value] += 1
            else:
                md5_dict[md5_value] = 1

            file_list.append({
                "目录层级": dir_levels,
                "文件名": filename,
                "大小(MB)": file_size_mb,
                "占比": size_percent,
                "格式": ext,
                "文件类型": get_file_type(ext),
                "创建日期": datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d'),
                "MD5": md5_value
            })

    # 标记重复文件
    for file_info in file_list:
        file_info["是否重复"] = md5_dict[file_info["MD5"]] > 1

    return file_list, max_depth + 1


def sort_file_list(file_list, max_depth):
    """多级目录排序函数"""

    def sort_key(item):
        dirs = item["目录层级"]
        dir_keys = [dirs[i].lower() if i < len(dirs) else "" for i in range(max_depth)]
        file_key = item["文件名"].lower()
        return tuple(dir_keys + [file_key])

    return sorted(file_list, key=sort_key)


def export_to_excel(folder_path, file_list, max_depth):
    # 先排序数据
    sorted_list = sort_file_list(file_list, max_depth)

    folder_name = os.path.basename(folder_path)
    timestamp = datetime.now().strftime("%Y%m%d")
    output_path = os.path.join(folder_path, f'{folder_name}-目录-{timestamp}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "文件目录"

    # 设置字体和颜色
    yahei_font = Font(name='微软雅黑', size=9)
    duplicate_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 生成表头（新增MD5和重复标记列）
    headers = [f"{i}级目录" for i in range(1, max_depth + 1)]
    headers += ["文件名", "大小(MB)", "占比(%)", "格式", "文件类型", "创建日期", "MD5", "重复文件"]
    ws.append(headers)

    # 写入数据
    for file_info in sorted_list:
        dir_levels = file_info["目录层级"]
        dir_data = [dir_levels[i] if i < len(dir_levels) else "" for i in range(max_depth)]

        row = dir_data + [
            file_info["文件名"],
            file_info["大小(MB)"],
            file_info["占比"],
            file_info["格式"],
            file_info["文件类型"],
            file_info["创建日期"],
            file_info["MD5"],
            "是" if file_info["是否重复"] else "否"
        ]
        ws.append(row)

    # 设置格式
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = yahei_font
            if cell.column == max_depth + 3:  # 设置百分比列
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            # 标记重复文件行
            if ws.cell(row=row_idx, column=len(headers)).value == "是":
                for cell in row:
                    cell.fill = duplicate_fill

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(headers) - 1)].width = 32  # MD5列
    ws.column_dimensions[openpyxl.utils.get_column_letter(len(headers))].width = 10  # 重复文件列

    wb.save(output_path)
    return output_path


def main():
    folder_path = select_folder()
    if not folder_path:
        print("未选择文件夹")
        return

    file_list, max_depth = get_file_info(folder_path)
    output_path = export_to_excel(folder_path, file_list, max_depth)
    print(f"目录已生成：{output_path}")


if __name__ == "__main__":
    main()
