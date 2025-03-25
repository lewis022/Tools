import csv
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import subprocess

# 隐藏Tkinter的主窗口
Tk().withdraw()

# 选择输入的CSV文件
input_csv_file = askopenfilename(title="选择要读取的CSV文件", filetypes=[("CSV Files", "*.csv")])
if not input_csv_file:
    print("未选择文件，程序退出。")
    exit()

# 自动生成输出路径
csv_dir = os.path.dirname(input_csv_file)  # 获取目录路径
base_name = os.path.splitext(os.path.basename(input_csv_file))[0]  # 获取文件名（不带扩展名）

# 输出合并后的XLSX文件路径
merged_xlsx_filename = f"{base_name}_output.xlsx"
output_xlsx_file = os.path.join(csv_dir, merged_xlsx_filename)  # 拼接路径

# 输出分列后的XLSX文件路径
split_xlsx_filename = f"{base_name}_split_output.xlsx"
split_output_xlsx_file = os.path.join(csv_dir, split_xlsx_filename)  # 拼接路径

# 读取CSV文件并将每一行合并成一个单元格
merged_rows = []
with open(input_csv_file, mode='r', encoding='utf-8') as csv_file:
    csv_reader = csv.reader(csv_file)
    for row in csv_reader:
        merged_cell = ''.join(row)  # 合并行中的所有单元格
        merged_rows.append([merged_cell])

# 输出合并后的内容到自动生成的XLSX文件
wb = Workbook()
ws = wb.active
for row in merged_rows:
    ws.append(row)
wb.save(output_xlsx_file)

# 读取生成的XLSX文件并按“;”分列
wb = load_workbook(output_xlsx_file)
ws = wb.active

# 创建新工作簿保存分列结果
new_wb = Workbook()
new_ws = new_wb.active

for row in ws.iter_rows(values_only=True):
    for cell in row:
        if cell:
            split_values = cell.split(';')  # 按“;”分列
            new_ws.append(split_values)

# 保存分列结果到新文件
new_wb.save(split_output_xlsx_file)

def open_file(path):
    """ 跨平台文件打开方法 """
    if os.name == "nt":  # Windows系统
        os.startfile(path)
    else:  # macOS/Linux系统
        subprocess.run(["open", path])  # macOS使用open命令

# 打开分列后的文件（新增函数调用）
open_file(split_output_xlsx_file)